#-*-coding:utf-8-*-
from openpyxl import load_workbook
from API_5.common import project_path
from API_5.common.read_config import ReadConfig
from API_5.common.my_log import MyLog
from API_5.common.get_data import GetData
class DoExcel:
    '''这是一个读取excel文件的类，并把测试数据写回excel
    file_name: 工作簿名
    sheet_name: 表单名'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.wb=load_workbook(self.file_name)

    def read_excel(self,section): #section配置文件里面的片段名，可以根据你的指定来执行具体的用例
        '''函数：完成excel表格数据的读取，有返回值
        section从配置文件里面控制读取那些用例'''
        sheet=self.wb[self.sheet_name] #打开表单
        test_data=[] #存储所有的数据至大列表中
        MyLog().info('开始读取数据了！')
        case_id=ReadConfig(project_path.conf_path,section,'case_id').get_str()
        tel_sheet=self.wb['tel'] #打开tel表单
        tel=tel_sheet.cell(2,1).value #获取tel值，为后面替换手机号做准备
        for i in range(2,sheet.max_row+1): #遍历每行测试用例
            row_data={} #读取每行数据以键值对的形式存储到字典里面
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Moudle']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url']=sheet.cell(i,4).value
            row_data['Method']=sheet.cell(i,5).value
            #判断并动态变化变化sql里面的值
            s_1=sheet.cell(i,7).value
            if 'mobile_phone' in str(s_1):
                row_data['Sql']=str(s_1).replace('mobile_phone',str(getattr(GetData,'MOBILEPHONE')))
            elif 'member_id' in str(s_1):
                row_data['Sql']=str(s_1).replace('member_id',str(getattr(GetData,'MEMBERID')))
            else:
                row_data['Sql']=sheet.cell(i,7).value
            row_data['ExceptedResult'] = sheet.cell(i,8).value
            # 判断并动态变化变化Params列表里面的值
            s=sheet.cell(i, 6).value #将参数赋值给字符串
            if 'tel' in str(s):#判断手机号值里面是否有tel字符串，有的话就替换掉，
                row_data['Params']=str(s).replace('tel',str(tel))
                tel+=1 #tel每替换一次后加1
                tel_sheet.cell(2,1).value=tel #加1后回写到tel表单的单元格
                self.wb.save(self.file_name) #保存
            elif 'mobile_phone' in str(s):
                row_data['Params']=str(s).replace('mobile_phone',str(getattr(GetData,'MOBILEPHONE')))
            elif 'member_id' in str(s):
                row_data['Params']=str(s).replace('member_id',str(getattr(GetData,'MEMBERID')))
            else: #判断不需要替换tel值的情况
                row_data['Params']=sheet.cell(i,6).value
            test_data.append(row_data) #添加每行数据至大列表中
        MyLog().info('读取数据结束！')
        self.wb.close()  # 读取完毕关闭工作簿
        final_data=[]
        if case_id=='all': #配置文件读取所有数据
            final_data=test_data
        else: #配置文件读取指定case_id数据
            for i in eval(case_id):
                final_data.append(test_data[i-1])
        return final_data #返回读取的测试数据

    def write_back(self,row,col,value):
        '''函数：完成写回数据至excel'''
        sheet=self.wb[self.sheet_name] #打开表单
        MyLog().info('开始写入值了！')
        try:
            sheet.cell(row,col).value=value #指定单元格进行赋值
            MyLog().info('写入完毕!')
        except Exception as e:
            MyLog().error('写入错误，错误信息：{}'.format(e))
        self.wb.save(self.file_name) #保存数据至工作簿
        self.wb.close() #关闭工作簿

#测试代码
if __name__=='__main__':
    de=DoExcel(project_path.case_path,'login')
    res=de.read_excel('LoginCASE')
    print(res)
    print(len(res))
    # de.write_back('test_case',2,1,1)










