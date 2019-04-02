from openpyxl import load_workbook
from API_4.common import project_path
from API_4.common.read_config_new import ReadConfig
from API_4.common.my_log import MyLog
#一般具有多个sheet表单的测试数据时不建议用这种方式，用API_3中的方法，分模块执行，这样报告也更条理清晰
#操作思路：一次性把需要测试的数据从excel表格中读取出来，到底取哪些用例根据配置文件来定
#配置文件的数据，这个时候是个字典：case_id={'register':[1,2,3,4,5],'login':[1,2,3,4,5],'recharge':[1,2,3,4,5,6,7]}
#设计的技巧，字典的key和表单名一致
#全部ddt，全部放到test_case里面，然后全部同时执行
class DoExcel:
    '''这是一个读取excel文件的类，并把测试数据写回excel
    file_name: 工作簿名
    sheet_name: 表单名'''
    def __init__(self,file_name):
        self.file_name=file_name
        self.wb=load_workbook(self.file_name)

    def read_excel(self): #section配置文件里面的片段名，可以根据你的指定来执行具体的用例
        '''函数：完成excel表格数据的读取，有返回值
        section从配置文件里面控制读取那些用例'''
        tel_sheet = self.wb['tel']  # 打开tel表单
        tel=tel_sheet.cell(2, 1).value  # 获取tel值，为后面替换手机号做准备
        case_id=ReadConfig(project_path.new_conf_path,'CASE','case_id').get_str()
        MyLog().info('开始读取数据了！')
        final_data=[]
        for key in eval(case_id):
            sheet_name=key #配置文件里面的字典key就是表单名
            self.sheet_name=key
            sheet=self.wb[sheet_name]
            test_data=[]  # 存储所有的数据至大列表中
            for i in range(2,sheet.max_row+1): #遍历每行测试用例
                row_data={} #读取每行数据以键值对的形式存储到字典里面
                row_data['CaseId']=sheet.cell(i,1).value
                row_data['Moudle']=sheet.cell(i,2).value
                row_data['Title']=sheet.cell(i,3).value
                row_data['Url']=sheet.cell(i,4).value
                row_data['Method']=sheet.cell(i,5).value
                row_data['ExceptedResult'] = sheet.cell(i, 7).value
                s=sheet.cell(i, 6).value #将参数赋值给字符串
                if 'tel' in str(s):#判断手机号值里面是否有tel字符串，有的话就替换掉，
                    row_data['Params']=str(s).replace('tel',str(tel))
                    tel+=1 #tel每替换一次后加1
                    tel_sheet.cell(2,1).value=tel #加1后回写到tel表单的单元格
                    self.wb.save(self.file_name) #保存
                else: #判断不需要替换tel值的情况
                    row_data['Params']=sheet.cell(i,6).value
                test_data.append(row_data) #添加每行数据至大列表中
            #根据配置文件筛选测试数据
            if eval(case_id)[key]=='all': #配置文件对应的字典值为all时读取所有数据
                final_data=test_data
            else: #否则根据配置文件读取指定数据
                for i in eval(case_id)[key]:
                    final_data.append(test_data[i-1])
            self.wb.close()  # 读取完毕关闭工作簿
        return final_data #返回读取的测试数据

    def write_back(self,sheet_name,row,col,value):
        '''函数：完成写回数据至excel'''
        sheet=self.wb[sheet_name] #打开表单
        MyLog().info('开始写入值了！')
        try:
            sheet.cell(row,col).value=value #指定单元格进行赋值
            MyLog().info('写入完毕!')
        except Exception as e:
            MyLog().error('写入错误，错误信息：{}'.format(e))
        self.wb.save(self.file_name) #保存数据至工作簿
        self.wb.close() #关闭工作簿

#测试代码
if __name__=='__main__':
    de=DoExcel(project_path.case_path)
    res=de.read_excel()
    print(res)
    print(len(res))
    # de.write_back('test_case',2,1,1)










