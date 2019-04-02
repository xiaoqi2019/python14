# 2）测试用例的数据存储在Excel中，并编写一个从Excel中读取数据的测试类，包含的函数能够读取测试数据，
# 并且能够写回测试结果，要求有返回值。
# 2）测试用例的数据存储在Excel中，并编写一个从Excel中读取数据的测试类，包含的函数能够读取测试数据，
# 并且能够写回测试结果，要求有返回值。
from openpyxl import load_workbook
from openpyxl import workbook
from API_1.common.read_config import ReadConfig
from API_1.common.my_log import MyLog
from API_1.common import project_path
class DoExcel:
    '''这是一个可以完成对excel文件进行读取测试数据和写回数据的类'''
    def __init__(self,file_name,sheet_name,tel_sheet_name):
        '''file_name: 读取数据的Excel工作簿文件名
        sheet_name: 读取数据的Excel表单名'''
        self.file__name=file_name
        self.sheet_name=sheet_name
        self.tel_sheet_name=tel_sheet_name
    def read_excel(self):
        '''完成对excel文件的读取，有返回值'''
        wb=load_workbook(self.file__name)
        sheet=wb[self.sheet_name]
        tel_sheet=wb[self.tel_sheet_name]
        tel_value=tel_sheet.cell(2,1).value
        MyLog().info('开始读取测试数据了')
        case_id=ReadConfig(project_path.conf_path,'CASE','case_id').get_str() #获取要执行用例的id
        test_data=[] #定义所有行的数据存到一个大列表里面
        for i in range(2,sheet.max_row+1):
            row_data={} #定义每行数据存在一个字列表里面
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Moudle']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url'] = sheet.cell(i,4).value
            row_data['Method']=sheet.cell(i,5).value
            row_data['ExceptedResult']=sheet.cell(i,7).value
            s = str(sheet.cell(i, 6).value)
            if 'tel' in s:
                row_data['Params'] = s.replace('tel',str(tel_value))
                tel_value+=1
                tel_sheet.cell(2, 1).value=tel_value
                wb.save(self.file__name)
            else:
                row_data['Params'] = s
            test_data.append(row_data)
        wb.close()
        final_data=[]#空列表 存储最终的测试用例数据
        if case_id=='all':#如果case_id==all 那就获取所有的用例数据
           final_data=test_data#把测试用例赋值给final_data这个变量
        else:#否则 如果是列表 那就获取列表里面指定id的用例的数据
           for i in eval(case_id):#遍历case_id 里面的值
               final_data.append(test_data[i-1])
        return final_data

    def write_back(self,row,col,value):
        '''完对excel文件的写入操作
        row: 写回数据的单元格的行值
        column: 写回数据单元格的列值
        value: 写回单元格的值'''
        wb=load_workbook(self.file__name)
        sheet=wb[self.sheet_name]
        MyLog().info('开始写入数据了！')
        try:
            sheet.cell(row,col).value=value #指定行列写入指定数据
            wb.save(self.file__name)  # 保存excel文件
            wb.close()
        except Exception as e:
            MyLog().error(e)
        MyLog().info('数据写入完毕！')


if __name__=='__main__':
    de=DoExcel(project_path.case_path,'test_case','tel')
    res=de.read_excel()
    print(res)
    # print(len(de.read_excel()))





































# from openpyxl import load_workbook
# from openpyxl import workbook
# from API_1.common.read_config import ReadConfig
# from API_1.common.my_log import MyLog
# from API_1.common import project_path
# class DoExcel:
#     '''这是一个可以完成对excel文件进行读取测试数据和写回数据的类'''
#     def __init__(self,file_name,sheet_name):
#         '''file_name: 读取数据的Excel工作簿文件名
#         sheet_name: 读取数据的Excel表单名'''
#         self.file__name=file_name
#         self.sheet_name=sheet_name
#         # self.tel_sheet_name=tel_sheet_name
#
#     def read_excel(self):
#         '''完成对excel文件的读取，有返回值'''
#         wb=load_workbook(self.file__name)
#         sheet=wb[self.sheet_name]
#         MyLog().info('开始读取测试数据了')
#         case_id=ReadConfig(project_path.conf_path,'CASE','case_id').get_str() #获取要执行用例的id
#         test_data=[] #定义所有行的数据存到一个大列表里面
#         for i in range(2,sheet.max_row+1):
#             row_data={} #定义每行数据存在一个字列表里面
#             row_data['CaseId']=sheet.cell(i,1).value
#             row_data['Moudle']=sheet.cell(i,2).value
#             row_data['Title']=sheet.cell(i,3).value
#             row_data['Url'] = sheet.cell(i,4).value
#             row_data['Method']=sheet.cell(i,5).value
#             row_data['ExceptedResult']=sheet.cell(i,7).value
#             row_data['Params'] = str(sheet.cell(i, 6).value)
#             test_data.append(row_data)
#         wb.close()
#         final_data=[]#空列表 存储最终的测试用例数据
#         if case_id=='all':#如果case_id==all 那就获取所有的用例数据
#            final_data=test_data#把测试用例赋值给final_data这个变量
#         else:#否则 如果是列表 那就获取列表里面指定id的用例的数据
#            for i in eval(case_id):#遍历case_id 里面的值
#                final_data.append(test_data[i-1])
#         return final_data
#
#     def write_back(self,row,col,value):
#         '''完对excel文件的写入操作
#         row: 写回数据的单元格的行值
#         column: 写回数据单元格的列值
#         value: 写回单元格的值'''
#         wb=load_workbook(self.file__name)
#         sheet=wb[self.sheet_name]
#         MyLog().info('开始写入数据了！')
#         try:
#             sheet.cell(row,col).value=value #指定行列写入指定数据
#             wb.save(self.file__name)  # 保存excel文件
#             wb.close()
#         except Exception as e:
#             MyLog().error(e)
#         MyLog().info('数据写入完毕！')
#
#
# if __name__=='__main__':
#     de=DoExcel(project_path.case_path,'test_case')
#     print(de.read_excel())
    # print(len(de.read_excel()))




