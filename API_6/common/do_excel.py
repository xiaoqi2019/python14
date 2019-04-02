#-*-coding:utf-8-*-
from openpyxl import load_workbook
from API_6.common import project_path
from API_6.common.read_config import ReadConfig
from API_6.common.my_log import MyLog
from API_6.common.get_data import GetData
import re
class DoExcel:
    '''这是一个读取excel文件的类，并把测试数据写回excel
    file_name: 工作簿名
    sheet_name: 表单名'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.wb=load_workbook(self.file_name)

    def read_excel(self,section): #section配置文件里面的片段名，可以根据你的指定来执行具体的用例
        '''函数：完成excel表格数据的读取，有返回值
        section从配置文件里面控制读取那些用例'''
        sheet=self.wb[self.sheet_name] #打开表单
        test_data=[] #存储所有的数据至大列表中
        MyLog().info('开始读取数据了！')
        case_id=ReadConfig(project_path.conf_path,section,'case_id').get_str()
        tel_sheet=self.wb['tel'] #打开tel表单
        tel=tel_sheet.cell(2,1).value #获取tel值，为后面替换手机号做准备
        for i in range(2,sheet.max_row+1): #遍历每行测试用例
            row_data={} #读取每行数据以键值对的形式存储到字典里面
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Moudle']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Url']=sheet.cell(i,4).value
            row_data['Method']=sheet.cell(i,5).value

            p='#(.*?)#'
            # 判断并动态变化变化Params列表里面的值
            s=sheet.cell(i, 6).value #将参数赋值给字符串
            #使用正则表达式判断是否存在需要替换的字符串
            if re.findall(p,s)!=None:
                for item in re.findall(p,s): #判断参数里面是否存在需要替换的字符串
                    if item=='tel':
                        s=re.sub(p,str(tel),s,count=1)
                        tel+=1
                        tel_sheet.cell(2,1).value=tel
                        self.wb.save(self.file_name)
                    if item=='mobile_phone':
                        s=re.sub(p,str(getattr(GetData,'MOBILEPHONE')),s,count=1)
                    if item=='mobile_pwd':
                        s=re.sub(p,str(getattr(GetData,'MEMBER_PWD')),s,count=1)
                    if item=='member_id':
                        s=re.sub(p,str(getattr(GetData,'MEMBERID')),s,count=1)
                row_data['Params']=s
            else: #判断不需要替换tel值的情况
                row_data['Params']=sheet.cell(i,6).value

            #判断并动态变化变化sql里面的值
            s_1=sheet.cell(i,7).value
            #判断sql语句存在的情况下
            if s_1:
                # 使用正则表达式判断是否存在需要替换的字符串
                if re.findall(p,s_1)!=None:
                    for j in re.findall(p,s_1):  # 判断参数里面是否存在需要替换的字符串
                        if j=='mobile_phone':
                            s_1=re.sub(p,str(getattr(GetData,'MOBILEPHONE')),s_1,count=1)
                        if j=='member_id':
                            s_1=re.sub(p,str(getattr(GetData,'MEMBERID')),s_1,count=1)
                    row_data['Sql']=s_1
            else:  # 判断不需要替换tel值的情况
                row_data['Sql']=sheet.cell(i,7).value

            row_data['ExceptedResult']=sheet.cell(i,8).value
            test_data.append(row_data) #添加每行数据至大列表中
        MyLog().info('读取数据结束！')
        self.wb.close()  # 读取完毕关闭工作簿
        final_data=[]
        if case_id=='all': #配置文件读取所有数据
            final_data=test_data
        else: #配置文件读取指定case_id数据
            for i in eval(case_id):
                final_data.append(test_data[i-1])
        return final_data #返回读取的测试数据

    def write_back(self,row,col,value):
        '''函数：完成写回数据至excel'''
        sheet=self.wb[self.sheet_name] #打开表单
        MyLog().info('开始写入值了！')
        try:
            sheet.cell(row,col).value=value #指定单元格进行赋值
            MyLog().info('写入完毕!')
        except Exception as e:
            MyLog().error('写入错误，错误信息：{}'.format(e))
        self.wb.save(self.file_name) #保存数据至工作簿
        self.wb.close() #关闭工作簿

#测试代码
if __name__=='__main__':
    de=DoExcel(project_path.case_path,'register')
    res=de.read_excel('RegisterCASE')
    print(res)
    print(len(res))

    de=DoExcel(project_path.case_path,'login')
    res=de.read_excel('LoginCASE')
    print(res)
    print(len(res))


    de=DoExcel(project_path.case_path,'recharge')
    res=de.read_excel('RechargeCASE')
    print(res)
    print(len(res))










