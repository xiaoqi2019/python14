#-*-coding:utf-8-*-
#@Time  :2019/2/26 18:53
#@Author:xiaoqi
#@File  :task_practice.py

#1：写一个类 类的作用是完成Excel数据的读写 新建表单的操作
# 函数一：读取指定表单的数据，
# 有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。
# 每一行都有 一个单独的row_list [[1,2,3],[4,5,6]] #每一行数据读取完毕后，把row_list存到大列表all_row_list
# 函数二：在指定的单元格写入指定的数据，并保存到当前Excel
# 函数三：新建一个Excel

from openpyxl import load_workbook
from openpyxl import workbook
from configparser import ConfigParser
from week_7.class_0225.read_config import ReadConfig
class DoExcel:
    '''这是一个类，完成Excel数据的读写 新建表单的操作'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_excel(self,button):
        '''函数一：读取所有的数据'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        all_data=[]
        if button==1:
            # 方法1：
            # for i in range(2,sheet.max_row+1):
            #     row_data=[]
            #     for j in range(1,sheet.max_column+1):
            #         res=sheet.cell(i,j).value
            #         row_data.append(res)
            #     all_data.append(row_data)
            # 方法2：
            for i in range(2,sheet.max_row+1):
                # row_data=[]
                # row_data.append(sheet.cell(i,1).value)
                # row_data.append(sheet.cell(i,2).value)
                # row_data.append(sheet.cell(i,3).value)
                # row_data.append(sheet.cell(i,4).value)
                # row_data.append(sheet.cell(i,5).value)
                # row_data.append(sheet.cell(i,6).value)
                # row_data.append(sheet.cell(i,7).value)
                # row_data.append(sheet.cell(i,8).value)
                #嵌套字典
                row_data={}
                row_data['用例编号'] = sheet.cell(i,1).value
                row_data['用例模块'] = sheet.cell(i,2).value
                row_data['用例标题'] = sheet.cell(i,3).value
                row_data['用例数据'] = sheet.cell(i,4).value
                row_data['用例步骤'] = sheet.cell(i,5).value
                row_data['期望结果'] = sheet.cell(i,6).value
                row_data['实际结果'] = sheet.cell(i,7).value
                row_data['测试结果'] = sheet.cell(i,8).value
                all_data.append(row_data)
        else:
            for i in eval(button):
                # row_data=[]
                # for j in range(1,sheet.max_column+1):
                #     res=sheet.cell(i+1,j).value
                #     row_data.append(res)
                # all_data.append(row_data)
                row_data={}
                row_data['用例编号'] = sheet.cell(i+1,1).value
                row_data['用例模块'] = sheet.cell(i+1,2).value
                row_data['用例标题'] = sheet.cell(i+1,3).value
                row_data['用例数据'] = sheet.cell(i+1,4).value
                row_data['用例步骤'] = sheet.cell(i+1,5).value
                row_data['期望结果'] = sheet.cell(i+1,6).value
                row_data['实际结果'] = sheet.cell(i+1,7).value
                row_data['测试结果'] = sheet.cell(i+1,8).value
                all_data.append(row_data)
        return all_data

if __name__=='__main__':
    button=ReadConfig('lemon.conf','TestCase','button').get_data()
    res=DoExcel('python_16.xlsx','Sheet1').read_excel(button)
    print(res)