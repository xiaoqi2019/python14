#3：结合日志类以及do_excel类，加上异常判断 与日志输出
from openpyxl import load_workbook
from openpyxl import workbook
from week_7.class_0225.read_config import ReadConfig
from week_7.class_0228.my_log import MyLog
class DoExcel:
    '''这是一个类，完成Excel数据的读写 新建表单的操作'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_excel(self,button):
        '''函数一：读取所有的数据'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        MyLog().info('开始读取测试数据了！')
        all_data=[]
        if button==1:
            for i in range(2,sheet.max_row+1):
                row_data=[]
                row_data.append(sheet.cell(i,1).value)
                row_data.append(sheet.cell(i,2).value)
                row_data.append(sheet.cell(i,3).value)
                row_data.append(sheet.cell(i,4).value)
                row_data.append(sheet.cell(i,5).value)
                row_data.append(sheet.cell(i,6).value)
                row_data.append(sheet.cell(i,7).value)
                row_data.append(sheet.cell(i,8).value)
                all_data.append(row_data)
            MyLog().info('测试数据读取完毕！')
        else:
            for i in eval(button):
                row_data=[]
                for j in range(1,sheet.max_column+1):
                    res=sheet.cell(i+1,j).value
                    row_data.append(res)
                all_data.append(row_data)
            MyLog().info('测试数据读取完毕！')
        return all_data

    def write_excel(self,row,column,value):
        '''在指定的单元格写入指定的数据，并保存到当前excel中'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        try:
            MyLog().info('开始写入数据了')
            sheet.cell(row,column).value=value
            wb.save(self.file_name)
            wb.close()
        except Exception as e:
            MyLog().error(e)
        MyLog().info('数据写入完毕！')

    def create_excel(self):
        '''函数三：新建一个Excel'''
        wb=workbook.Workbook()
        wb.create_sheet(self.sheet_name)
        wb.save(self.file_name)

if __name__=='__main__':
    button=ReadConfig('lemon.conf','TestCase','button').get_data()
    de=DoExcel('python_16.xlsx','sheet3')
    res_1=de.read_excel(button)
    print(res_1)
    res_2=de.write_excel(2,1,'TestId')
    de.create_excel()
