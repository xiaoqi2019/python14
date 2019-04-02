#-*-coding:utf-8-*-
#@Time  :2019/2/25 14:53
#@Author:xiaoqi
#@File  :learn_excel.py
#python使用openpyxl模块来操作Excel
#为什么操作excel
#禅道 xmind excel
#模块openpyxl里面函数load_workbook是打开excel文件的
#workbook工作簿 sheet表单 cell单元格
#注意学习时类比open函数，操作excle，以xlsx结尾的才可以

#打开工作簿
from openpyxl import load_workbook
# wb=load_workbook('python_14.xlsx')
wb=load_workbook('C:\\Users\朱泽剑\PycharmProjects\python14\week_6\class_0222\python_14.xlsx')
# #定位表单
sheet=wb['Sheet']
# 定位单元格，默认先行后列
value=sheet.cell(3,1).value
# 方法2：指定行列值，从1开始数的
res=sheet.cell(row=3,column=1).value
print(res)

# 把第五行的数据全部读出来
# for i in (1,2,3,4,5):
#     res=sheet.cell(5,i).value
#     print(res)

# 得到最大行数和列数的函数
# print(sheet.max_row) #最大行数
# print(sheet.max_column) #最大列数

#读取所有行所有列的数据
# for i in range(1,sheet.max_row+1):
#     for j in range(1,sheet.max_column+1):
#         res=sheet.cell(i,j).value
#         if res!=None:
#             print(res)

# 写数据
# 方法1：
# sheet.cell(1,1,'作者：小七')#行，列value
# 方法2：
# sheet.cell(1,1).value='作者：小琪琪'
# wb.save('python_15.xlsx')

# 新建一个工作簿
# from openpyxl import workbook
# wb=workbook.Workbook()
# wb.save('python_16.xlsx')
















