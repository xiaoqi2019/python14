#-*-coding:utf-8-*-
#@Time  :2019/2/25 18:12
#@Author:xiaoqi
#@File  :task_02.py

# 2：思考：分别将我们学过的数据类型  int float boolean str list tuple dict
# 写到每个单元格里面，观察，你通过openpyxl操作后拿到的数据分别是是什么类型。
from openpyxl import load_workbook
wb=load_workbook('python_16.xlsx')
sheet=wb['Sheet1']
for i in range(1,sheet.max_row+1):
    res=sheet.cell(i,1).value
    print(res)
    print(type(res))
# value=sheet.cell(row,column).value
