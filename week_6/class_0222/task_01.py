#-*-coding:utf-8-*-
#@Time  :2019/2/25 16:44
#@Author:xiaoqi
#@File  :task_01.py

#1：写一个类 类的作用是完成Excel数据的读写 新建表单的操作
# 函数一：读取指定表单的数据，
  # 有一个列表row_list,把每一行的每一个单元格的数据存到row_list里面去。
  # 每一行都有 一个单独的row_list [[1,2,3],[4,5,6]] #每一行数据读取完毕后，把row_list存到大列表all_row_list
# 函数二：在指定的单元格写入指定的数据，并保存到当前Excel
# 函数三：新建一个Excel

from openpyxl import load_workbook
from openpyxl import workbook
class DoExcel:
    '''这是一个类，完成Excel数据的读写 新建表单的操作'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name

    def read_excel(self):
        '''函数一：读取所有的数据，以嵌套列表的形似存储，每一行都是一个子列表，
        每个单元格都是子列表的一个元素'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        #方法一：(嵌套列表)
        all_row_list=[]#所有的测试数据放在一个大的列表里面
        for i in range(2,sheet.max_row+1):#测试数据从第二行开始，第一行是标题
            row_list=[]#每一行数据存在一个子列表里面
            for j in range(1,sheet.max_column+1):
                res=sheet.cell(i,j).value
                row_list.append(res)
            all_row_list.append(row_list)
        return all_row_list#返回大列表

        #方法二：(嵌套字典)
        # test_data=[]
        # for i in range(2,sheet.max_row+1):
        #     row_dict={}#每一行数据存在一个字典里面
        #     row_dict['用例编号']=sheet.cell(i,1).value
        #     row_dict['测试项目']=sheet.cell(i,2).value
        #     row_dict['测试标题']=sheet.cell(i,3).value
        #     row_dict['测试数据']=sheet.cell(i,4).value
        #     row_dict['操作步骤']=sheet.cell(i,5).value
        #     row_dict['期望结果']=sheet.cell(i,6).value
        #     row_dict['实际结果']=sheet.cell(i,7).value
        #     row_dict['测试结果']=sheet.cell(i,8).value
        #     test_data.append(row_dict)
        # return test_data


    def write_excel(self,row,column,value):
        '''函数二：在指定的单元格写入指定的数据，并保存到当前Excel'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        sheet.cell(row,column).value=value
        wb.save(self.file_name)
        wb.close()#每次写操作完要关闭掉


    def create_excel(self):
        '''函数三：新建一个Excel'''
        wb=workbook.Workbook()
        wb.create_sheet(self.sheet_name)
        wb.save(self.file_name)

if __name__=='__main__':
    d=DoExcel('python_14.xlsx','Sheet')
    res=d.read_excel()
    print(res)
    # d.write_excel(1,1,'TestId')
    # d.create_excel('test')

