from openpyxl import load_workbook
from openpyxl import workbook
from week_8.class_0307.read_config import ReadConfig
from week_8.class_0307.my_log import MyLog
class DoExcel:
    '''这是一个可以完成excel文件读写，新建操作的类'''
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def read_excel(self,button):
        '''函数：完成用例数据的读取，按照嵌套列表的形式输出'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        MyLog().info('开始读取测试数据了！')
        all_data=[] #读取所有行用例
        if button==1:
            for i in range(2,sheet.max_row+1):
                row_date=[] #读取每行用例
                for j in range(1,sheet.max_column+1):
                    row_date.append(sheet.cell(i,j).value)
                all_data.append(row_date)
            MyLog().info('数据读入完毕！')
        else:
            for i in button:
                row_date=[]
                for j in range(1,sheet.max_column+1):
                    row_date.append(sheet.cell(i+1,j).value)
                all_data.append(row_date)
            MyLog().info('数据读入完毕！')
        return all_data
    def write_excel(self,row,column,value):
        '''函数：完成excel表格里面的数据的写入操作'''
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        MyLog().info('开始写入数据了！')
        try:
            sheet.cell(row,column).value=value #将指定数据写入指定的行和列
            wb.save(self.file_name)  # 保存文件
            wb.close()  # 关闭文件
        except Exception as e:
            MyLog().error(e)
        MyLog().info('数据写入完毕！')

    def create_excel(self):
        wb=workbook.Workbook()
        wb.create_sheet(self.sheet_name)
        wb.save(self.file_name)


if __name__=='__main__':
    button=ReadConfig('xiaoqi.conf','TestCase','button').get_data()
    de=DoExcel('python_14.xlsx','Sheet1')
    value=de.read_excel(button)
    print(value)
    # de.write_excel(1,1,'用例编号')
